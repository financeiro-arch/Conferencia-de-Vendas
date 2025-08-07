import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Conferência de Vendas - Grupo Óticas Visão", page_icon="📊")

st.title("📋 Conferência de Vendas")
st.subheader("Grupo Óticas Visão")

st.markdown("---")
st.markdown("Envie as duas planilhas para validar as vendas com base nos campos:")
st.markdown("- Código NSU\n- Código de Autorização\n- Código da Venda\n- Data\n- Valor\n- Loja")

uploaded_file1 = st.file_uploader("📎 Enviar Planilha 1 (Ex: PagSeguro ou REDE)", type=["csv", "xlsx"], key="file1")
uploaded_file2 = st.file_uploader("📎 Enviar Planilha 2 (Ex: Extrato de Vendas)", type=["csv", "xlsx"], key="file2")

if uploaded_file1 and uploaded_file2:
    def read_file(uploaded_file):
        if uploaded_file.name.endswith('.xlsx'):
            return pd.read_excel(uploaded_file, dtype=str)
        else:
            return pd.read_csv(uploaded_file, dtype=str)

    def normalizar_colunas(df):
        df.columns = df.columns.str.strip().str.lower()
        renomear = {}
        for col in df.columns:
            if "nsu" in col:
                renomear[col] = "codigo nsu"
            elif "autoriz" in col:
                renomear[col] = "codigo de autorizacao"
            elif "venda" in col and "codigo" in col:
                renomear[col] = "codigo da venda"
            elif any(term in col for term in ["data", "emissao"]):
                renomear[col] = "data"
            elif any(term in col for term in ["valor", "bruto"]):
                renomear[col] = "valor"
            elif any(term in col for term in ["loja", "local"]):
                renomear[col] = "loja"
        return df.rename(columns=renomear)

    try:
        df1 = read_file(uploaded_file1)
        df2 = read_file(uploaded_file2)

        st.markdown("### 📄 Pré-visualização das planilhas")
        with st.expander("📘 Planilha 1"):
            st.dataframe(df1.head())
        with st.expander("📗 Planilha 2"):
            st.dataframe(df2.head())

        df1 = normalizar_colunas(df1)
        df2 = normalizar_colunas(df2)

        chaves = ["codigo nsu", "codigo de autorizacao", "codigo da venda", "data", "valor", "loja"]

        if all(col in df1.columns for col in chaves) and all(col in df2.columns for col in chaves):

            df2['Status Conferência'] = df2.apply(
                lambda row: "Conferido" if ((df1[chaves] == row[chaves]).all(axis=1).any()) else "Erro",
                axis=1
            )

            st.success("✅ Conferência finalizada!")
            st.dataframe(df2)

            export = st.radio("Exportar resultado como:", ["Excel com cores"], horizontal=True)

            if export == "Excel com cores":
                output = BytesIO()
                df2.to_excel(output, index=False, sheet_name="Resultado")
                output.seek(0)

                wb = load_workbook(filename=output)
                ws = wb.active

                status_col_idx = None
                for idx, cell in enumerate(ws[1], start=1):
                    if cell.value == "Status Conferência":
                        status_col_idx = idx
                        break

                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
                    status_cell = row[status_col_idx - 1]
                    if status_cell.value == "Conferido":
                        for cell in row:
                            cell.fill = green_fill
                    elif status_cell.value == "Erro":
                        for cell in row:
                            cell.fill = red_fill

                final_output = BytesIO()
                wb.save(final_output)
                final_output.seek(0)

                st.download_button(
                    "⬇️ Baixar Resultado em Excel",
                    final_output,
                    file_name="resultado_conferencia.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        else:
            st.error("❌ As colunas obrigatórias não foram encontradas em ambas as planilhas. Mesmo após tentativa de normalização.")
            st.code("\n".join(chaves))

    except Exception as e:
        st.error(f"Erro ao processar os arquivos: {e}")

else:
    st.info("🔄 Aguarde o envio das duas planilhas para iniciar a conferência.")
